"""
app.py — Orchestrateur principal (léger)
Toutes les routes métier sont dans modules/
"""
from flask import (Flask, render_template, request, jsonify,
                   redirect, url_for, session, send_file, flash)
import hashlib, io
from datetime import datetime, date
from functools import wraps

# ── DB & helpers ────────────────────────────────────────────
from database import get_db, init_db
from modules.helpers import (login_required, get_current_user,
                              get_param, calculer_penalites, gen_num, annees_non_payees)

# ── Blueprints ───────────────────────────────────────────────
from modules.config         import bp as config_bp
from modules.contribuables  import bp as ctb_bp
from modules.tnb            import bp as tnb_bp
from modules.tdb            import bp as tdb_bp
from modules.stationnement  import bp as sta_bp
from modules.fourriere      import bp as fou_bp
from modules.occupation     import bp as odp_bp
from modules.location       import bp as loc_bp
from modules.souks          import bp as sou_bp

# ── Application ──────────────────────────────────────────────
app = Flask(__name__)
app.secret_key = 'fiscalite_communale_secret_2024'

for bp in (config_bp, ctb_bp, tnb_bp, tdb_bp, sta_bp, fou_bp, odp_bp, loc_bp, sou_bp):
    app.register_blueprint(bp)

DB = 'fiscalite.db'


# ════════════════════════════════════════════════════════════
#  AUTH
# ════════════════════════════════════════════════════════════
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        pwd = hashlib.sha256(request.form['password'].encode()).hexdigest()
        conn = get_db()
        user = conn.execute('SELECT * FROM utilisateurs WHERE email=? AND mot_de_passe=? AND actif=1',
                            (request.form['email'], pwd)).fetchone()
        conn.close()
        if user:
            session['user_id'] = user['id']
            return redirect(url_for('index'))
        return render_template('login.html', error='Email ou mot de passe incorrect')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


# ════════════════════════════════════════════════════════════
#  DASHBOARD
# ════════════════════════════════════════════════════════════
@app.route('/')
@login_required
def index():
    user = get_current_user()
    conn = get_db()
    stats = {
        'contribuables':    conn.execute('SELECT COUNT(*) as c FROM contribuables WHERE actif=1').fetchone()['c'],
        'terrains':         conn.execute('SELECT COUNT(*) as c FROM terrains WHERE actif=1').fetchone()['c'],
        'bulletins_attente':conn.execute("SELECT COUNT(*) as c FROM bulletins WHERE statut='en_attente'").fetchone()['c'],
        'avis_emis':        conn.execute("SELECT COUNT(*) as c FROM avis_non_paiement WHERE statut='emis'").fetchone()['c'],
        'total_emis':       conn.execute("SELECT COALESCE(SUM(montant_total),0) as s FROM declarations").fetchone()['s'],
        'total_paye':       conn.execute("SELECT COALESCE(SUM(montant),0) as s FROM bulletins WHERE statut='paye'").fetchone()['s'],
    }
    recentes = conn.execute('''SELECT d.*, c.nom, c.prenom, c.raison_sociale FROM declarations d
        JOIN contribuables c ON d.contribuable_id=c.id ORDER BY d.date_creation DESC LIMIT 8''').fetchall()
    conn.close()
    return render_template('index.html', user=user, stats=stats, recentes=recentes)


# ════════════════════════════════════════════════════════════
#  DÉCLARATIONS (communes à tous les modules)
# ════════════════════════════════════════════════════════════
@app.route('/declarations/creer', methods=['POST'])
@login_required
def creer_declaration():
    user = get_current_user()
    f = request.form
    module     = f['module']
    ref_id     = int(f['reference_id'])
    contrib_id = int(f['contribuable_id'])
    annee      = int(f.get('annee', datetime.now().year))
    base       = float(f.get('base_calcul', 0))
    taux       = float(f.get('taux', 0))
    principal  = round(base * taux / 100 if taux else base, 2)
    date_ech   = f.get('date_echeance', '')
    date_decl  = f.get('date_declaration', date.today().isoformat())
    hors_delai = f.get('hors_delai') == '1'
    penalite, majoration, amende = 0, 0, 0
    if hors_delai:
        a_pct  = get_param(module, 'AMENDE_NON_DECLARATION', 15)
        amende = max(round(principal * a_pct / 100, 2), 500)
    if date_ech and date_decl > date_ech:
        penalite, majoration = calculer_penalites(principal, date_ech, date_decl, module)
    total = round(principal + penalite + majoration + amende, 2)
    if total < 200:
        total = 0; statut = 'sous_seuil'
    else:
        statut = 'emis'
    num = gen_num('DCL', 'declarations')
    conn = get_db()
    conn.execute('''INSERT INTO declarations
        (numero,module,reference_id,contribuable_id,commune_id,annee,trimestre,
         base_calcul,taux,montant_principal,penalite_retard,majoration,amende_non_declaration,montant_total,
         statut,date_declaration,date_echeance,agent_id,notes)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
        (num, module, ref_id, contrib_id, 1, annee, int(f.get('trimestre', 0)),
         base, taux, principal, penalite, majoration, amende, total, statut,
         date_decl, date_ech, user['id'], f.get('notes', '')))
    conn.commit(); conn.close()
    flash(f'Déclaration {num} — Total: {total:.2f} DH ✅', 'success')
    routes_map = {
        'TNB':               'tnb.tnb_paiement',
        'DEBITS_BOISSONS':   'tdb.tdb_paiement',
        'STATIONNEMENT':     'sta.sta_paiement',
        'OCCUPATION_DOMAINE':'odp.odp_paiement',
        'FOURRIERE':         'fou.fou_paiement',
        'LOCATION_LOCAUX':   'loc.loc_paiement',
        'AFFERMAGE_SOUKS':   'sou.sou_paiement',
    }
    if module in routes_map:
        return redirect(url_for(routes_map[module], id=ref_id))
    return redirect(url_for('paiements'))


# ════════════════════════════════════════════════════════════
#  BULLETINS / PAIEMENTS
# ════════════════════════════════════════════════════════════
@app.route('/paiements')
@login_required
def paiements():
    user = get_current_user()
    conn = get_db()
    items = conn.execute('''SELECT b.*, d.module, d.annee, c.nom, c.prenom, c.raison_sociale
        FROM bulletins b JOIN declarations d ON b.declaration_id=d.id
        JOIN contribuables c ON b.contribuable_id=c.id ORDER BY b.date_creation DESC''').fetchall()
    decls_sans_bulletin = conn.execute('''SELECT d.*, c.nom, c.prenom, c.raison_sociale FROM declarations d
        JOIN contribuables c ON d.contribuable_id=c.id
        WHERE d.statut="emis" AND d.montant_total>0
        AND d.id NOT IN (SELECT declaration_id FROM bulletins WHERE statut IN ("en_attente","paye"))
        ORDER BY d.date_creation DESC''').fetchall()
    conn.close()
    return render_template('paiements/paiements.html', user=user, items=items,
                           decls=decls_sans_bulletin, today=date.today().isoformat())

@app.route('/bulletins/creer', methods=['POST'])
@login_required
def creer_bulletin():
    user = get_current_user()
    if not user['peut_creer_bulletin']:
        flash('Accès refusé', 'danger')
        return redirect(url_for('paiements'))
    f = request.form
    conn = get_db()
    decl = conn.execute('SELECT * FROM declarations WHERE id=?', (f['declaration_id'],)).fetchone()
    if decl:
        num = gen_num('BUL', 'bulletins', 'numero_bulletin')
        conn.execute('''INSERT INTO bulletins (numero_bulletin,declaration_id,contribuable_id,commune_id,montant,mode_paiement,date_paiement,agent_id,notes)
            VALUES (?,?,?,?,?,?,?,?,?)''',
            (num, decl['id'], decl['contribuable_id'], decl['commune_id'], decl['montant_total'],
             f.get('mode_paiement','especes'), f.get('date_paiement', date.today().isoformat()),
             user['id'], f.get('notes','')))
        conn.commit()
        flash(f'Bulletin {num} créé — En attente validation régisseur ✅', 'success')
    conn.close()
    return redirect(url_for('paiements'))

@app.route('/bulletins/<int:id>/valider', methods=['POST'])
@login_required
def valider_bulletin(id):
    user = get_current_user()
    if not user['peut_valider_paiement']:
        flash('Accès refusé — Réservé au Régisseur', 'danger')
        return redirect(url_for('paiements'))
    f = request.form
    num_quittance = f.get('numero_quittance', '').strip()
    date_quittance = f.get('date_quittance', date.today().isoformat())
    if not num_quittance:
        flash('Le numéro de quittance est obligatoire', 'danger')
        return redirect(url_for('paiements'))
    conn = get_db()
    b = conn.execute('SELECT * FROM bulletins WHERE id=?', (id,)).fetchone()
    if b:
        # Valider CE bulletin
        conn.execute("""UPDATE bulletins 
            SET statut='paye', regisseur_id=?, numero_quittance=?, date_quittance=?
            WHERE id=?""", (user['id'], num_quittance, date_quittance, id))
        conn.execute("UPDATE declarations SET statut='paye', date_paiement=? WHERE id=?",
                     (date_quittance, b['declaration_id']))
        # Valider aussi tous les autres bulletins avec le même numero_bulletin
        # (cas multi-trimestres TDB créés avec le même N° de BV)
        autres = conn.execute(
            "SELECT id, declaration_id FROM bulletins WHERE numero_bulletin=? AND id!=? AND statut='en_attente'",
            (b['numero_bulletin'], id)).fetchall()
        for ab in autres:
            conn.execute("UPDATE bulletins SET statut='paye', regisseur_id=?, numero_quittance=?, date_quittance=? WHERE id=?",
                         (user['id'], num_quittance, date_quittance, ab['id']))
            conn.execute("UPDATE declarations SET statut='paye', date_paiement=? WHERE id=?",
                         (date_quittance, ab['declaration_id']))
        conn.commit()
        total_val = 1 + len(autres)
        flash(f'✅ Paiement validé — Quittance N° {num_quittance} — {total_val} déclaration(s) soldée(s)', 'success')
    conn.close()
    return redirect(url_for('paiements'))


@app.route('/bulletins/<int:id>/rejeter', methods=['POST'])
@login_required
def rejeter_bulletin(id):
    user = get_current_user()
    if not user['peut_valider_paiement']:
        flash('Accès refusé', 'danger')
        return redirect(url_for('paiements'))
    f = request.form
    motif = f.get('motif_rejet', 'Non précisé').strip()
    conn = get_db()
    b = conn.execute('SELECT * FROM bulletins WHERE id=?', (id,)).fetchone()
    if b:
        conn.execute("UPDATE bulletins SET statut='rejete', motif_rejet=?, regisseur_id=? WHERE id=?",
                     (motif, user['id'], id))
        conn.execute("UPDATE declarations SET statut='emis' WHERE id=?", (b['declaration_id'],))
        conn.commit()
        flash(f'❌ Bulletin N° {b["numero_bulletin"]} rejeté : {motif}', 'danger')
    conn.close()
    return redirect(url_for('paiements'))

@app.route('/bulletins/valider-masse', methods=['POST'])
@login_required
def valider_bulletins_masse():
    user = get_current_user()
    if not user['peut_valider_paiement']:
        flash('Accès refusé — Réservé au Régisseur', 'danger')
        return redirect(url_for('paiements'))
    f = request.form
    num_quittance = f.get('numero_quittance', '').strip()
    date_quittance = f.get('date_quittance', date.today().isoformat())
    bulletin_ids = f.getlist('bulletin_ids')
    if not num_quittance:
        flash('Le numéro de quittance est obligatoire', 'danger')
        return redirect(url_for('paiements'))
    if not bulletin_ids:
        flash('Aucun bulletin sélectionné', 'warn')
        return redirect(url_for('paiements'))
    conn = get_db()
    count = 0
    for bid in bulletin_ids:
        b = conn.execute("SELECT * FROM bulletins WHERE id=? AND statut='en_attente'", (bid,)).fetchone()
        if b:
            conn.execute("""UPDATE bulletins 
                SET statut='paye', regisseur_id=?, numero_quittance=?, date_quittance=?
                WHERE id=?""", (user['id'], num_quittance, date_quittance, int(bid)))
            conn.execute("UPDATE declarations SET statut='paye', date_paiement=? WHERE id=?",
                         (date_quittance, b['declaration_id']))
            count += 1
    conn.commit()
    conn.close()
    flash(f'✅ {count} bulletin(s) validé(s) — Quittance N° {num_quittance}', 'success')
    return redirect(url_for('paiements'))


# ════════════════════════════════════════════════════════════
#  AVIS D'IMPOSITION
# ════════════════════════════════════════════════════════════
@app.route('/avis')
@login_required
def avis():
    user = get_current_user()
    conn = get_db()
    items = conn.execute('''SELECT a.*, d.module, d.annee, c.nom, c.prenom, c.raison_sociale, c.adresse
        FROM avis_non_paiement a JOIN declarations d ON a.declaration_id=d.id
        JOIN contribuables c ON a.contribuable_id=c.id ORDER BY a.date_emission DESC''').fetchall()
    decls = conn.execute('''SELECT d.*, c.nom, c.prenom, c.raison_sociale FROM declarations d
        JOIN contribuables c ON d.contribuable_id=c.id
        WHERE d.statut="emis" AND d.montant_total>0
        AND d.id NOT IN (SELECT declaration_id FROM avis_non_paiement WHERE statut="emis")''').fetchall()
    conn.close()
    return render_template('admin/avis.html', user=user, items=items, decls=decls)

@app.route('/avis/generer', methods=['POST'])
@login_required
def generer_avis():
    conn = get_db()
    mode   = request.form.get('mode', 'individuel')
    lot_id = f"LOT{datetime.now().strftime('%Y%m%d%H%M%S')}"
    if mode == 'lot':
        decls = conn.execute('''SELECT * FROM declarations WHERE statut="emis" AND montant_total>0
            AND id NOT IN (SELECT declaration_id FROM avis_non_paiement WHERE statut="emis")''').fetchall()
        for d in decls:
            num = gen_num('AVS', 'avis_non_paiement', 'numero_avis')
            conn.execute('''INSERT INTO avis_non_paiement (numero_avis,declaration_id,contribuable_id,commune_id,montant_du,date_emission,lot_id)
                VALUES (?,?,?,?,?,?,?)''',
                (num, d['id'], d['contribuable_id'], d['commune_id'], d['montant_total'], date.today().isoformat(), lot_id))
    else:
        decl_id = request.form.get('declaration_id')
        if decl_id:
            d = conn.execute('SELECT * FROM declarations WHERE id=?', (decl_id,)).fetchone()
            if d:
                num = gen_num('AVS', 'avis_non_paiement', 'numero_avis')
                conn.execute('''INSERT INTO avis_non_paiement (numero_avis,declaration_id,contribuable_id,commune_id,montant_du,date_emission)
                    VALUES (?,?,?,?,?,?)''',
                    (num, d['id'], d['contribuable_id'], d['commune_id'], d['montant_total'], date.today().isoformat()))
    conn.commit(); conn.close()
    flash('Avis générés ✅', 'success')
    return redirect(url_for('avis'))


# ════════════════════════════════════════════════════════════
#  UTILISATEURS & COMMUNES
# ════════════════════════════════════════════════════════════
@app.route('/utilisateurs')
@login_required
def utilisateurs():
    user = get_current_user()
    conn = get_db()
    items = conn.execute('''SELECT u.*, r.nom as role_nom FROM utilisateurs u
        JOIN roles r ON u.role_id=r.id WHERE u.actif=1''').fetchall()
    roles = conn.execute('SELECT * FROM roles').fetchall()
    communes = conn.execute('SELECT * FROM communes WHERE actif=1').fetchall()
    conn.close()
    return render_template('admin/utilisateurs.html', user=user, items=items, roles=roles, communes=communes)

@app.route('/utilisateurs/ajouter', methods=['POST'])
@login_required
def ajouter_utilisateur():
    f = request.form
    pwd = hashlib.sha256(f['mot_de_passe'].encode()).hexdigest()
    conn = get_db()
    conn.execute('INSERT OR IGNORE INTO utilisateurs (nom,prenom,email,mot_de_passe,role_id,commune_id) VALUES (?,?,?,?,?,?)',
        (f['nom'], f['prenom'], f['email'], pwd, f['role_id'], f.get('commune_id', 1)))
    conn.commit(); conn.close()
    flash('Utilisateur ajouté ✅', 'success')
    return redirect(url_for('utilisateurs'))

@app.route('/utilisateurs/<int:id>/modifier', methods=['POST'])
@login_required
def modifier_utilisateur(id):
    f = request.form
    conn = get_db()
    if f.get('mot_de_passe'):
        pwd = hashlib.sha256(f['mot_de_passe'].encode()).hexdigest()
        conn.execute('UPDATE utilisateurs SET nom=?,prenom=?,email=?,mot_de_passe=?,role_id=?,commune_id=? WHERE id=?',
            (f['nom'], f['prenom'], f['email'], pwd, f['role_id'], f.get('commune_id', 1), id))
    else:
        conn.execute('UPDATE utilisateurs SET nom=?,prenom=?,email=?,role_id=?,commune_id=? WHERE id=?',
            (f['nom'], f['prenom'], f['email'], f['role_id'], f.get('commune_id', 1), id))
    conn.commit(); conn.close()
    flash('Utilisateur modifié ✅', 'success')
    return redirect(url_for('utilisateurs'))

@app.route('/utilisateurs/<int:id>/supprimer', methods=['POST'])
@login_required
def supprimer_utilisateur(id):
    conn = get_db()
    conn.execute('UPDATE utilisateurs SET actif=0 WHERE id=?', (id,))
    conn.commit(); conn.close()
    return redirect(url_for('utilisateurs'))

@app.route('/communes')
@login_required
def communes():
    user = get_current_user()
    conn = get_db()
    items = conn.execute('SELECT * FROM communes WHERE actif=1').fetchall()
    conn.close()
    return render_template('admin/communes.html', user=user, items=items)

@app.route('/communes/ajouter', methods=['POST'])
@login_required
def ajouter_commune():
    f = request.form
    conn = get_db()
    conn.execute('INSERT OR IGNORE INTO communes (nom,nom_ar,region,province,code) VALUES (?,?,?,?,?)',
        (f['nom'], f.get('nom_ar',''), f['region'], f['province'], f['code']))
    conn.commit(); conn.close()
    return redirect(url_for('communes'))


# ════════════════════════════════════════════════════════════
#  EXPORT EXCEL
# ════════════════════════════════════════════════════════════
@app.route('/export/<module>/excel')
@login_required
def export_excel(module):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    conn = get_db()
    data = conn.execute('''SELECT d.numero, c.nom||" "||COALESCE(c.prenom,"") as contribuable, c.cin,
        d.annee, d.base_calcul, d.taux, d.montant_principal, d.penalite_retard, d.majoration,
        d.amende_non_declaration, d.montant_total, d.statut, d.date_declaration
        FROM declarations d JOIN contribuables c ON d.contribuable_id=c.id
        WHERE d.module=? ORDER BY d.date_creation DESC''', (module,)).fetchall()
    conn.close()
    wb = Workbook(); ws = wb.active; ws.title = module[:31]
    hf = Font(bold=True, color='FFFFFF')
    hfill = PatternFill('solid', fgColor='1e3a5f')
    hdrs = ['N° Décl.','Contribuable','CIN','Année','Base','Taux%','Principal',
            'Pénalité','Majoration','Amende','TOTAL','Statut','Date']
    for i, h in enumerate(hdrs, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = hf; cell.fill = hfill
        ws.column_dimensions[cell.column_letter].width = 15
    for r, row in enumerate(data, 2):
        for i, v in enumerate(row, 1):
            ws.cell(row=r, column=i, value=v)
    out = io.BytesIO(); wb.save(out); out.seek(0)
    return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'{module}_{datetime.now():%Y%m%d}.xlsx')


# ════════════════════════════════════════════════════════════
#  API JSON
# ════════════════════════════════════════════════════════════
@app.route('/api/contribuable/<int:id>')
@login_required
def api_contribuable(id):
    conn = get_db()
    c = conn.execute('SELECT * FROM contribuables WHERE id=?', (id,)).fetchone()
    conn.close()
    return jsonify(dict(c) if c else {})

@app.route('/api/calcul')
@login_required
def api_calcul():
    montant    = float(request.args.get('montant', 0))
    ech        = request.args.get('echeance', '')
    pay        = request.args.get('paiement', date.today().isoformat())
    module     = request.args.get('module', 'GLOBAL')
    hors_delai = request.args.get('hors_delai', '0') == '1'
    p, m = calculer_penalites(montant, ech, pay, module)
    amende = 0
    if hors_delai:
        a_pct  = get_param(module, 'AMENDE_NON_DECLARATION', 15)
        amende = max(round(montant * a_pct / 100, 2), 500)
    return jsonify({'penalite': p, 'majoration': m, 'amende': amende,
                    'total': round(montant + p + m + amende, 2)})

@app.route('/api/tarifs/<module>')
@login_required
def api_tarifs(module):
    conn = get_db()
    today = date.today().isoformat()
    tarifs = conn.execute('''SELECT t.* FROM tarifs t
        JOIN rubriques r ON t.rubrique_id=r.id
        WHERE r.module=? AND t.actif=1
          AND t.date_debut <= ?
          AND (t.date_fin IS NULL OR t.date_fin >= ?)
        ORDER BY t.libelle''', (module, today, today)).fetchall()
    conn.close()
    return jsonify([dict(t) for t in tarifs])

@app.route('/api/stats')
@login_required
def api_stats():
    conn = get_db()
    modules = ['TNB','DEBITS_BOISSONS','TRANSPORT_VOYAGEURS','STATIONNEMENT',
               'OCCUPATION_DOMAINE','FOURRIERE','LOCATION_LOCAUX','AFFERMAGE_SOUKS']
    result = {}
    for m in modules:
        r = conn.execute("SELECT COUNT(*) as c, COALESCE(SUM(montant_total),0) as t FROM declarations WHERE module=?", (m,)).fetchone()
        result[m] = {'count': r['c'], 'total': round(r['t'], 2)}
    conn.close()
    return jsonify(result)


# ════════════════════════════════════════════════════════════
#  POINT D'ENTRÉE
# ════════════════════════════════════════════════════════════
if __name__ == '__main__':
    init_db()
    import socket
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80)); ip = s.getsockname()[0]; s.close()
    except:
        ip = 'localhost'
    print(f"\n{'='*55}\n  JIBAYAT — Gestion Fiscale Communale\n  Local : http://localhost:5000\n  Réseau: http://{ip}:5000\n  Login : admin@commune.ma / admin123\n{'='*55}\n")
    app.run(host='0.0.0.0', port=5000, debug=False)
